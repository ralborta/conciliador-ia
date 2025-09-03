import pandas as pd
import logging
import re
import unicodedata
import math
import traceback
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path
from datetime import datetime
import uuid
import os

logger = logging.getLogger(__name__)

def s(x):
    """Helper para convertir cualquier valor a string de forma segura"""
    if x is None: 
        return ""
    if isinstance(x, float) and math.isnan(x): 
        return ""
    return str(x)

def safe_join(*parts: object) -> str:
    """Helper para concatenar strings de forma segura"""
    return "".join(s(p) for p in parts)

def strip_xls_float(sv: str) -> str:
    """Quita .0 heredado de Excel sin romper strings"""
    return sv[:-2] if isinstance(sv, str) and sv.endswith(".0") else sv

class ClienteProcessor:
    def __init__(self):
        self.tipo_doc_mapping = {
            '80': 'CUIT',
            '96': 'DNI',
            'CUIT': 'CUIT',
            'DNI': 'DNI'
        }
        
        # Mapeo de condiciones IVA a abreviaciones
        self.condicion_iva_mapping = {
            'Monotributista': 'MT',
            'Responsable Inscripto': 'RI', 
            'Consumidor Final': 'CF',
            'Exento': 'EX',
            'MT': 'MT',
            'RI': 'RI',
            'CF': 'CF',
            'EX': 'EX'
        }
        
        # Base de datos de prefijos CUIT por provincia
        self.prefijos_provincia = {
            '20': 'Buenos Aires',
            '21': 'Buenos Aires',
            '22': 'Chaco',
            '23': 'Chubut',
            '24': 'C√≥rdoba',
            '25': 'Corrientes',
            '26': 'Entre R√≠os',
            '27': 'Formosa',
            '28': 'Jujuy',
            '29': 'La Pampa',
            '30': 'La Rioja',
            '31': 'Mendoza',
            '32': 'Misiones',
            '33': 'Neuqu√©n',
            '34': 'R√≠o Negro',
            '35': 'Salta',
            '36': 'San Juan',
            '37': 'San Luis',
            '38': 'Santa Cruz',
            '39': 'Santa Fe',
            '40': 'Santiago del Estero',
            '41': 'Tierra del Fuego',
            '42': 'Tucum√°n'
        }
        
        # Base de datos de c√≥digos postales por rangos de DNI
        self.cp_rangos_dni = {
            # Buenos Aires Capital
            '10': 'Buenos Aires Capital',
            '11': 'Buenos Aires Capital',
            '12': 'Buenos Aires Capital',
            '13': 'Buenos Aires Capital',
            '14': 'Buenos Aires Capital',
            '15': 'Buenos Aires Capital',
            '16': 'Buenos Aires Capital',
            '17': 'Buenos Aires Capital',
            '18': 'Buenos Aires Capital',
            '19': 'Buenos Aires Capital',
            
            # Buenos Aires GBA
            '20': 'Buenos Aires GBA',
            '21': 'Buenos Aires GBA',
            '22': 'Buenos Aires GBA',
            '23': 'Buenos Aires GBA',
            '24': 'Buenos Aires GBA',
            '25': 'Buenos Aires GBA',
            '26': 'Buenos Aires GBA',
            '27': 'Buenos Aires GBA',
            '28': 'Buenos Aires GBA',
            '29': 'Buenos Aires GBA',
            
            # C√≥rdoba
            '50': 'C√≥rdoba Capital',
            '51': 'C√≥rdoba Capital',
            '52': 'C√≥rdoba Capital',
            '53': 'C√≥rdoba Capital',
            '54': 'C√≥rdoba Capital',
            '55': 'C√≥rdoba Capital',
            '56': 'C√≥rdoba Capital',
            '57': 'C√≥rdoba Capital',
            '58': 'C√≥rdoba Capital',
            '59': 'C√≥rdoba Capital',
            
            # Santa Fe
            '30': 'Santa Fe Capital',
            '31': 'Santa Fe Capital',
            '32': 'Santa Fe Capital',
            '33': 'Santa Fe Capital',
            '34': 'Santa Fe Capital',
            '35': 'Santa Fe Capital',
            '36': 'Santa Fe Capital',
            '37': 'Santa Fe Capital',
            '38': 'Santa Fe Capital',
            '39': 'Santa Fe Capital',
            
            # Mendoza
            '55': 'Mendoza Capital',
            '56': 'Mendoza Capital',
            '57': 'Mendoza Capital',
            '58': 'Mendoza Capital',
            '59': 'Mendoza Capital',
            
            # Tucum√°n
            '40': 'Tucum√°n Capital',
            '41': 'Tucum√°n Capital',
            '42': 'Tucum√°n Capital',
            '43': 'Tucum√°n Capital',
            '44': 'Tucum√°n Capital',
            '45': 'Tucum√°n Capital',
            '46': 'Tucum√°n Capital',
            '47': 'Tucum√°n Capital',
            '48': 'Tucum√°n Capital',
            '49': 'Tucum√°n Capital'
        }
    
    def normalizar_texto(self, texto: str) -> str:
        """Normaliza texto eliminando caracteres especiales y normalizando espacios"""
        if pd.isna(texto) or texto is None:
            return ""
        
        texto = str(texto).strip()
        
        # Normalizar caracteres especiales
        texto = unicodedata.normalize('NFD', texto)
        texto = re.sub(r'[^\w\s]', ' ', texto)
        
        # Reemplazar m√∫ltiples espacios por uno solo
        texto = re.sub(r'\s+', ' ', texto)
        
        return texto.upper()
    
    def normalizar_identificador(self, identificador: str) -> str:
        """Normaliza identificadores (CUIT/DNI) eliminando separadores"""
        if pd.isna(identificador) or identificador is None:
            return ""
        
        identificador = str(identificador).strip()
        # Eliminar todos los caracteres no num√©ricos
        identificador = re.sub(r'[^\d]', '', identificador)
        return identificador
    
    def validar_y_formatear_dni(self, dni: str) -> Tuple[bool, str]:
        """Valida y formatea DNI a 8 d√≠gitos"""
        dni_limpio = self.normalizar_identificador(dni)
        
        if len(dni_limpio) == 7:
            dni_limpio = '0' + dni_limpio
        elif len(dni_limpio) == 8:
            pass
        else:
            return False, dni
        
        return True, dni_limpio
    
    def validar_y_formatear_cuit(self, cuit: str) -> Tuple[bool, str]:
        """Valida y formatea CUIT con guiones"""
        cuit_limpio = self.normalizar_identificador(cuit)
        
        if len(cuit_limpio) != 11:
            return False, cuit
        
        # Formatear con guiones XX-XXXXXXXX-X
        cuit_formateado = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10:]}"
        
        # Validar checksum b√°sico (implementaci√≥n simplificada)
        # En producci√≥n, implementar validaci√≥n completa de CUIT
        return True, cuit_formateado
    
    def mapear_tipo_documento(self, codigo: str) -> Optional[str]:
        """Mapea c√≥digo AFIP a tipo de documento"""
        codigo_limpio = str(codigo).strip()
        return self.tipo_doc_mapping.get(codigo_limpio)
    
    def determinar_condicion_iva(self, tipo_doc: str, numero_doc: str) -> str:
        """Determina condici√≥n IVA basada en tipo y n√∫mero de documento"""
        if tipo_doc == "DNI":
            return "CF"  # Consumidor Final
        elif tipo_doc == "CUIT":
            # L√≥gica simplificada - en producci√≥n usar reglas m√°s complejas
            if numero_doc.startswith('20') or numero_doc.startswith('23') or numero_doc.startswith('24'):
                return "RI"  # Responsable Inscripto
            else:
                return "MT"  # Monotributista
        return "CF"  # Consumidor Final por defecto
    
    def convertir_condicion_iva_a_abreviacion(self, condicion_iva: str) -> str:
        """Convierte cualquier condici√≥n IVA a su abreviaci√≥n de dos letras"""
        condicion_limpia = str(condicion_iva).strip()
        return self.condicion_iva_mapping.get(condicion_limpia, "CF")  # CF por defecto
    
    def obtener_provincia_por_cuit(self, cuit: str) -> str:
        """Obtiene provincia por prefijo de CUIT"""
        if not cuit or len(cuit) < 2:
            return ""
        
        # Limpiar CUIT y tomar primeros 2 d√≠gitos
        cuit_limpio = self.normalizar_identificador(cuit)
        if len(cuit_limpio) >= 2:
            prefijo = cuit_limpio[:2]
            return self.prefijos_provincia.get(prefijo, "")
        
        return ""
    
    def obtener_localidad_por_dni(self, dni: str) -> str:
        """Obtiene localidad por primeros d√≠gitos del DNI"""
        if not dni or len(dni) < 2:
            return ""
        
        # Limpiar DNI y tomar primeros 2 d√≠gitos
        dni_limpio = self.normalizar_identificador(dni)
        if len(dni_limpio) >= 2:
            prefijo = dni_limpio[:2]
            return self.cp_rangos_dni.get(prefijo, "")
        
        return ""
    
    def obtener_provincia_por_dni(self, dni: str, df_historico: pd.DataFrame = None) -> str:
        """Obtiene provincia por DNI usando datos hist√≥ricos"""
        if not dni or df_historico is None or df_historico.empty:
            return ""
        
        dni_limpio = self.normalizar_identificador(dni)
        
        # Buscar en datos hist√≥ricos por DNI
        for _, row in df_historico.iterrows():
            # Buscar columna de DNI
            dni_col = None
            for col in df_historico.columns:
                if any(keyword in col.lower() for keyword in ['dni', 'documento', 'identificador', 'numeroidentificacion']):
                    dni_col = col
                    break
            
            if dni_col:
                dni_historico = self.normalizar_identificador(str(row[dni_col]))
                if dni_limpio == dni_historico:
                    # Buscar columna de provincia
                    provincia_col = None
                    for col in df_historico.columns:
                        if any(keyword in col.lower() for keyword in ['provincia', 'prov', 'estado']):
                            provincia_col = col
                            break
                    
                    if provincia_col:
                        provincia = str(row[provincia_col]).strip()
                        if provincia and provincia.lower() not in ['nan', 'none', '']:
                            return provincia
        
        return ""
    
    def obtener_provincia_por_nombre(self, nombre: str, df_historico: pd.DataFrame = None) -> str:
        """Intenta obtener provincia por nombre del cliente"""
        if not nombre or df_historico is None or df_historico.empty:
            return ""
        
        nombre_normalizado = self.normalizar_texto(nombre)
        
        # Buscar coincidencias exactas o similares en datos hist√≥ricos
        for _, row in df_historico.iterrows():
            # Buscar columna de nombre
            nombre_col = None
            for col in df_historico.columns:
                if any(keyword in col.lower() for keyword in ['nombre', 'razon', 'cliente']):
                    nombre_col = col
                    break
            
            if nombre_col:
                nombre_historico = self.normalizar_texto(str(row[nombre_col]))
                if nombre_normalizado == nombre_historico:
                    # Buscar columna de provincia
                    provincia_col = None
                    for col in df_historico.columns:
                        if any(keyword in col.lower() for keyword in ['provincia', 'prov', 'estado']):
                            provincia_col = col
                            break
                    
                    if provincia_col:
                        provincia = str(row[provincia_col]).strip()
                        if provincia and provincia.lower() not in ['nan', 'none', '']:
                            return provincia
        
        return ""
    
    def detectar_nuevos_clientes(
        self,
        df_portal: pd.DataFrame,
        df_xubio: pd.DataFrame,
        df_cliente: Optional[pd.DataFrame] = None
    ) -> Tuple[List[Dict], List[Dict]]:
        """Detecta clientes nuevos comparando portal vs Xubio"""
        
        try:
            # Debug: Log de columnas disponibles
            logger.info(f"Columnas del Portal: {list(df_portal.columns)}")
            logger.info(f"Columnas de Xubio: {list(df_xubio.columns)}")
            if df_cliente is not None:
                logger.info(f"Columnas del Cliente: {list(df_cliente.columns)}")
            
            nuevos_clientes = []
            errores = []
            
            # Normalizar maestros
            xubio_identificadores = set()
            xubio_nombres = set()
        
            for _, row in df_xubio.iterrows():
                # Buscar columnas de identificador - Mapeo espec√≠fico para Xubio
                id_cols = [col for col in df_xubio.columns if any(keyword in col.lower() 
                            for keyword in ['cuit', 'dni', 'documento', 'identificador', 'numeroidentificacion', 'numero_identificacion', 'numeroidentificacion', 'numeroidentificacion'])]
                
                if id_cols:
                    identificador = self.normalizar_identificador(str(row[id_cols[0]]))
                    if identificador:
                        xubio_identificadores.add(identificador)
                
                # Buscar columnas de nombre - Mapeo espec√≠fico para Xubio
                nombre_cols = [col for col in df_xubio.columns if any(keyword in col.lower() 
                              for keyword in ['nombre', 'razon', 'cliente', 'NOMBRE'])]
                
                if nombre_cols:
                    nombre = self.normalizar_texto(str(row[nombre_cols[0]]))
                    if nombre:
                        xubio_nombres.add(nombre)
            
            # Procesar cada fila del portal
            for idx, row in df_portal.iterrows():
                try:
                    # Buscar columnas relevantes - Mapeo m√°s flexible para archivos del portal
                    tipo_doc_col = self._encontrar_columna(df_portal.columns, ['tipo_doc', 'tipo_documento', 'tipo', 'ct_kind0f', 'TIPO_DOC', 'Tipo Doc. Comprador'])
                    numero_doc_col = self._encontrar_columna(df_portal.columns, ['NUMERO_DOC', 'numero_doc', 'Numero de Documento', 'numero de documento', 'numero_documento', 'nro. doc. comprador', 'nro doc comprador', 'nro. doc comprador', 'dni', 'cuit', 'CUIT', 'NUMERO_DOC'])
                    nombre_col = self._encontrar_columna(df_portal.columns, ['nombre', 'razon_social', 'cliente', 'NOMBRE', 'denominaci√É¬≥n comprador', 'denominacion comprador', 'denominaci√£¬≥n comprador', 'denominaci√≥n comprador'])
                    
                    # DEBUG: Verificar qu√© columnas se encontraron
                    fila_num = idx[0] if isinstance(idx, tuple) else idx
                    logger.info(f"Fila {fila_num + 1}: tipo_doc_col='{tipo_doc_col}', numero_doc_col='{numero_doc_col}', nombre_col='{nombre_col}'")
                    
                    # MOSTRAR COLUMNAS DISPONIBLES EN PRIMERA FILA
                    if fila_num == 0:
                        logger.info(f"üìã COLUMNAS DISPONIBLES: {list(df_portal.columns)}")
                        logger.info(f"üîç B√öSQUEDA DE COLUMNAS:")
                        logger.info(f"   - Tipo documento: '{tipo_doc_col}'")
                        logger.info(f"   - N√∫mero documento: '{numero_doc_col}'")
                        logger.info(f"   - Nombre: '{nombre_col}'")
                    
                    # NO FORZAR COLUMNAS - USAR LAS DETECTADAS
                    # if tipo_doc_col != 'Tipo Doc. Comprador':
                    #     logger.warning(f"Fila {fila_num + 1}: Cambiando tipo_doc_col de '{tipo_doc_col}' a 'Tipo Doc. Comprador'")
                    #     tipo_doc_col = 'Tipo Doc. Comprador'
                    
                    if not all([tipo_doc_col, numero_doc_col, nombre_col]):
                        errores.append({
                            'origen_fila': safe_join("Portal fila ", fila_num + 1),
                            'tipo_error': 'Columnas faltantes',
                            'detalle': safe_join('No se encontraron columnas requeridas. Disponibles: ', ', '.join(df_portal.columns), '. Buscadas: tipo_doc=', bool(tipo_doc_col), ', numero_doc=', bool(numero_doc_col), ', nombre=', bool(nombre_col)),
                            'valor_original': str(row.to_dict())
                        })
                        continue
                    
                    # Extraer valores
                    tipo_doc_codigo = str(row[tipo_doc_col]).strip()
                    numero_doc = str(row[numero_doc_col]).strip()
                    nombre = str(row[nombre_col]).strip()
                    
                    # Mapear tipo de documento
                    tipo_documento = self.mapear_tipo_documento(tipo_doc_codigo)
                    if not tipo_documento:
                        logger.warning(f"Fila {fila_num + 1}: C√≥digo de documento no reconocido: '{tipo_doc_codigo}'")
                        errores.append({
                            'origen_fila': safe_join("Portal fila ", fila_num + 1),
                            'tipo_error': 'Tipo de documento no reconocido',
                            'detalle': safe_join('C√≥digo ', tipo_doc_codigo, ' no mapeable. C√≥digos v√°lidos: 80=CUIT, 96=DNI'),
                            'valor_original': tipo_doc_codigo
                        })
                        continue
                    else:
                        logger.info(f"Fila {fila_num + 1}: Conversi√≥n exitosa '{tipo_doc_codigo}' ‚Üí '{tipo_documento}'")
                    
                    # Validar y formatear documento
                    if tipo_documento == "DNI":
                        valido, numero_formateado = self.validar_y_formatear_dni(numero_doc)
                        logger.info(f"Fila {fila_num + 1}: Validaci√≥n DNI '{numero_doc}' ‚Üí '{numero_formateado}' (v√°lido: {valido})")
                    else:  # CUIT
                        valido, numero_formateado = self.validar_y_formatear_cuit(numero_doc)
                        logger.info(f"Fila {fila_num + 1}: Validaci√≥n CUIT '{numero_doc}' ‚Üí '{numero_formateado}' (v√°lido: {valido})")
                
                    if not valido:
                        errores.append({
                            'origen_fila': safe_join("Portal fila ", fila_num + 1),
                            'tipo_error': safe_join(tipo_documento, ' inv√°lido'),
                            'detalle': 'Longitud o formato incorrecto',
                            'valor_original': numero_doc
                        })
                        continue
                    
                    # Verificar si es cliente nuevo
                    identificador_normalizado = self.normalizar_identificador(numero_doc)
                    nombre_normalizado = self.normalizar_texto(nombre)
                    
                    logger.info(f"Fila {fila_num + 1}: Cliente '{nombre}' (Doc: {numero_formateado}) - Normalizado: '{identificador_normalizado}'")
                    
                    if identificador_normalizado in xubio_identificadores:
                        logger.info(f"Fila {fila_num + 1}: Cliente ya existe en Xubio - Saltando")
                        continue  # Cliente ya existe en Xubio
                    else:
                        logger.info(f"Fila {fila_num + 1}: Cliente NUEVO detectado - Procesando")
                    
                    # Buscar provincia - PRIMERO intentar por documento en Xubio
                    provincia = self._obtener_provincia_por_documento(numero_formateado, df_xubio)
                    if provincia:
                        logger.info(f"Fila {fila_num + 1}: Provincia desde Xubio: {provincia}")
                
                    # Si no se encuentra, usar m√©todo anterior como fallback
                    if not provincia:
                        provincia = self._buscar_provincia(row, df_portal.columns, df_cliente)
                        if provincia:
                            logger.info(f"Fila {fila_num + 1}: Provincia desde datos del portal: {provincia}")
                    
                    # Si a√∫n no se encuentra, intentar por prefijo CUIT o DNI
                    if not provincia and tipo_documento == "CUIT":
                        provincia = self.obtener_provincia_por_cuit(numero_formateado)
                        if provincia:
                            logger.info(f"Fila {fila_num + 1}: Provincia determinada por prefijo CUIT: {provincia}")
                    
                    if not provincia and tipo_documento == "DNI":
                        # Primero intentar por datos hist√≥ricos
                        if df_cliente is not None:
                            provincia = self.obtener_provincia_por_dni(numero_formateado, df_cliente)
                            if provincia:
                                logger.info(f"Fila {fila_num + 1}: Provincia determinada por DNI en datos hist√≥ricos: {provincia}")
                        
                        # Si no se encontr√≥, intentar por rangos de DNI (c√≥digos postales)
                        if not provincia:
                            provincia = self.obtener_localidad_por_dni(numero_formateado)
                            if provincia:
                                logger.info(f"Fila {fila_num + 1}: Provincia determinada por rango DNI: {provincia}")
                    
                    # Si a√∫n no se encuentra, marcar como sin provincia
                    if not provincia:
                        provincia = ""  # Sin provincia
                        logger.warning(f"Fila {fila_num + 1}: No se pudo determinar provincia")
                    
                    # Determinar condici√≥n IVA
                    condicion_iva = self.determinar_condicion_iva(tipo_documento, numero_formateado)
                    logger.info(f"Fila {fila_num + 1}: Condici√≥n IVA determinada: {condicion_iva}")
                
                    # Determinar localidad
                    localidad = ""
                    if tipo_documento == "DNI":
                        localidad = self.obtener_localidad_por_dni(numero_formateado)
                        if localidad:
                            logger.info(f"Fila {fila_num + 1}: Localidad determinada por DNI: {localidad}")
                    
                    # Crear cliente nuevo
                    nuevo_cliente = {
                        'nombre': nombre,
                        'tipo_documento': tipo_documento,
                        'numero_documento': numero_formateado,
                        'condicion_iva': condicion_iva,
                        'provincia': provincia,
                        'localidad': localidad,
                        'cuenta_contable': 'Deudores por ventas'
                    }
                    
                    nuevos_clientes.append(nuevo_cliente)
                    logger.info(f"Fila {fila_num + 1}: ‚úÖ CLIENTE CREADO: {nombre} ({tipo_documento}: {numero_formateado}) - {provincia}")
                
                except Exception as e:
                    errores.append({
                        'origen_fila': safe_join("Portal fila ", fila_num + 1),
                        'tipo_error': 'Error de procesamiento',
                        'detalle': str(e),
                        'valor_original': str(row.to_dict())
                    })
        
            # Eliminar duplicados por identificador
            clientes_unicos = self._eliminar_duplicados(nuevos_clientes)
            
            return clientes_unicos, errores
            
        except Exception as e:
            logger.error("IMPORT FAIL\n%s", traceback.format_exc())
            return [], [{
                'origen_fila': 'Sistema',
                'tipo_error': f'{type(e).__name__}',
                'detalle': str(e),
                'valor_original': 'Error en procesamiento general'
            }]
    
    def _encontrar_columna(self, columnas: List[str], keywords: List[str]) -> Optional[str]:
        """Encuentra columna que contenga alguno de los keywords"""
        # Ordenar keywords por especificidad (m√°s espec√≠ficos primero)
        keywords_ordenados = sorted(keywords, key=len, reverse=True)
        
        for col in columnas:
            col_lower = col.lower()
            for keyword in keywords_ordenados:
                if keyword in col_lower:
                    return col
        return None
    
    def _buscar_provincia(
        self, 
        row: pd.Series, 
        columnas_portal: List[str], 
        df_cliente: Optional[pd.DataFrame]
    ) -> Optional[str]:
        """Busca provincia en el orden: Portal -> Excel Cliente"""
        
        # Buscar en portal
        provincia_col = self._encontrar_columna(columnas_portal, ['provincia', 'prov'])
        if provincia_col and pd.notna(row[provincia_col]):
            return str(row[provincia_col]).strip()
        
        # Buscar en excel del cliente si est√° disponible
        if df_cliente is not None:
            # Buscar por nombre del cliente
            nombre_col_portal = self._encontrar_columna(columnas_portal, ['nombre', 'razon_social', 'cliente'])
            if nombre_col_portal:
                nombre_cliente = str(row[nombre_col_portal]).strip()
                
                # Buscar en excel del cliente - Mapeo espec√≠fico para archivo del cliente
                nombre_col_cliente = self._encontrar_columna(df_cliente.columns, ['nombre', 'razon_social', 'cliente', 'RAZON SOCIAL / APELLIDO', 'NOMBRE'])
                provincia_col_cliente = self._encontrar_columna(df_cliente.columns, ['provincia', 'prov', 'Provincia / Estado / Region'])
                
                if nombre_col_cliente and provincia_col_cliente:
                    for _, cliente_row in df_cliente.iterrows():
                        if self.normalizar_texto(str(cliente_row[nombre_col_cliente])) == self.normalizar_texto(nombre_cliente):
                            provincia = str(cliente_row[provincia_col_cliente]).strip()
                            if provincia and provincia.lower() not in ['nan', 'none', '']:
                                return provincia
        
        return None
    
    def _obtener_provincia_por_documento(
        self, 
        numero_documento: str, 
        df_xubio: pd.DataFrame
    ) -> Optional[str]:
        """Recupera provincia desde Xubio usando n√∫mero de documento como clave"""
        
        try:
            # Normalizar n√∫mero del portal (agregar guiones para formato Xubio)
            if len(numero_documento) == 11:  # CUIT sin guiones
                numero_normalizado = f"{numero_documento[:2]}-{numero_documento[2:10]}-{numero_documento[10:]}"
            else:
                numero_normalizado = numero_documento
            
            # Buscar en Xubio por n√∫mero de documento
            cliente_xubio = df_xubio[df_xubio['Numero de Documento'] == numero_normalizado]
            
            if not cliente_xubio.empty:
                # Buscar columna de provincia
                provincia_col = None
                for col in df_xubio.columns:
                    if 'provincia' in col.lower() or 'estado' in col.lower() or 'region' in col.lower():
                        provincia_col = col
                        break
                
                if provincia_col:
                    provincia = str(cliente_xubio.iloc[0][provincia_col]).strip()
                    if provincia and provincia.lower() not in ['nan', 'none', '']:
                        return provincia
            
            return None
            
        except Exception as e:
            logger.warning(f"Error al buscar provincia por documento {numero_documento}: {e}")
            return None
    
    def _eliminar_duplicados(self, clientes: List[Dict]) -> List[Dict]:
        """Elimina duplicados bas√°ndose en n√∫mero de documento"""
        vistos = set()
        unicos = []
        
        for cliente in clientes:
            identificador = cliente['numero_documento']
            if identificador not in vistos:
                vistos.add(identificador)
                unicos.append(cliente)
        
        return unicos
    
    def generar_archivo_importacion(
        self, 
        clientes: List[Dict], 
        output_dir: Path,
        cuenta_contable_default: str = "Deudores por ventas"
    ) -> str:
        """Genera archivo CSV para importaci√≥n en Xubio"""
        
        output_dir.mkdir(parents=True, exist_ok=True)

        # Estructura exacta seg√∫n la imagen del archivo de clientes
        columnas_xubio = [
            "NUMERODECONTROL",
            "NOMBRE", 
            "CODIGO",
            "TIPOIDE",
            "NUMEROIDENTIF",
            "CONDICI",
            "EMAIL",
            "TELEFON",
            "DIRECCI",
            "PROVINCIA",
            "LOCALID",
            "CUENTA",
            "LISTADE",
            "OBSER",
            "CIONES"
        ]
        
        # Crea DF vac√≠o con columnas si df_nuevos viene vac√≠o
        if not clientes or len(clientes) == 0:
            df_out = pd.DataFrame(columns=columnas_xubio)
        else:
            df_out = self._mapear_a_xubio(clientes, cuenta_contable_default, columnas_xubio)

        # Asegurar que solo tengamos las columnas esperadas
        df_out = df_out[columnas_xubio]
        
        # Reemplazar NaN por cadenas vac√≠as en todas las columnas
        for col in df_out.columns:
            df_out[col] = df_out[col].fillna("")
        
        # Asegurar que todos los campos vac√≠os sean cadenas vac√≠as (no NaN)
        for col in df_out.columns:
            if col not in ["NUMERODECONTROL", "NOMBRE", "TIPOIDE", "NUMEROIDENTIF", "CONDICI", "PROVINCIA", "CUENTA"]:
                df_out[col] = [""] * len(df_out)

        nombre = f"clientes_xubio_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
        ruta = Path(output_dir) / nombre

        # Generar archivo Excel usando openpyxl directamente para controlar el formato
        from openpyxl import Workbook
        from openpyxl.styles import Font
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes Xubio"
        
        # Escribir headers
        for col_idx, col_name in enumerate(columnas_xubio, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Escribir datos
        for row_idx, cliente in enumerate(clientes, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)  # NUMERODECONTROL
            ws.cell(row=row_idx, column=2, value=cliente.get("nombre", ""))
            ws.cell(row=row_idx, column=3, value=" ")  # CODIGO con espacio
            ws.cell(row=row_idx, column=4, value=cliente.get("tipo_documento", "DNI"))
            ws.cell(row=row_idx, column=5, value=cliente.get("numero_documento", ""))
            ws.cell(row=row_idx, column=6, value=cliente.get("condicion_iva", "CF"))
            ws.cell(row=row_idx, column=7, value=" ")  # EMAIL con espacio
            ws.cell(row=row_idx, column=8, value=" ")  # TELEFON con espacio
            ws.cell(row=row_idx, column=9, value=" ")  # DIRECCI con espacio
            ws.cell(row=row_idx, column=10, value=cliente.get("provincia", ""))
            ws.cell(row=row_idx, column=11, value=cliente.get("localidad", " ") if cliente.get("localidad") else " ")
            ws.cell(row=row_idx, column=12, value=cuenta_contable_default)
            ws.cell(row=row_idx, column=13, value=" ")  # LISTADE con espacio
            ws.cell(row=row_idx, column=14, value=" ")  # OBSER con espacio
            ws.cell(row=row_idx, column=15, value=" ")  # CIONES con espacio
        
        wb.save(ruta)
        logger.info(f"Archivo Excel de importaci√≥n generado: {ruta}")
        return str(ruta)
    
    def generar_reporte_errores(
        self, 
        errores: List[Dict], 
        output_dir: Path
    ) -> str:
        """Genera reporte CSV de errores"""
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        if not errores:
            # Crear reporte vac√≠o con headers
            df_err = pd.DataFrame(columns=['origen_fila', 'tipo_error', 'detalle', 'valor_original'])
        else:
            df_err = pd.DataFrame(errores)
        
        nombre = f"reporte_errores_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.csv"
        ruta = Path(output_dir) / nombre
        
        # UTF-8 con BOM p/Excel
        df_err.to_csv(ruta, index=False, encoding="utf-8-sig")
        logger.info(f"Reporte de errores generado: {ruta}")
        return str(ruta)

    def _mapear_a_xubio(self, clientes: List[Dict], cuenta_contable_default: str, columnas_xubio: list) -> pd.DataFrame:
        """Mapea los datos de clientes al formato exacto de la imagen"""
        if not clientes:
            return pd.DataFrame(columns=columnas_xubio)
        
        # Crear DataFrame con todas las columnas inicializadas con cadenas vac√≠as
        out = pd.DataFrame(index=range(len(clientes)))
        
        # Inicializar todas las columnas con None (pandas lo maneja mejor)
        for col in columnas_xubio:
            out[col] = [None] * len(clientes)
        
        # NUMERODECONTROL - N√∫mero secuencial
        out["NUMERODECONTROL"] = list(range(1, len(clientes) + 1))
        
        # NOMBRE - Nombre del cliente
        for cand in ["nombre", "razon_social", "cliente", "denominacion"]:
            if cand in clientes[0]:
                out["NOMBRE"] = [cliente.get(cand, "") for cliente in clientes]
                break
        if "NOMBRE" not in out:
            out["NOMBRE"] = [cliente.get("nombre", "") for cliente in clientes]

        # TIPOIDE - Tipo de documento
        out["TIPOIDE"] = [cliente.get("tipo_documento", "DNI") for cliente in clientes]

        # NUMEROIDENTIF - N√∫mero de documento
        out["NUMEROIDENTIF"] = [cliente.get("numero_documento", "") for cliente in clientes]

        # CONDICI - Condici√≥n IVA (abreviaci√≥n de 2 letras)
        if "condicion_iva" in clientes[0]:
            out["CONDICI"] = [cliente.get("condicion_iva", "CF") for cliente in clientes]
        else:
            out["CONDICI"] = ["CF"] * len(clientes)

        # PROVINCIA - Provincia del cliente
        out["PROVINCIA"] = [cliente.get("provincia", "") for cliente in clientes]

        # LOCALID - Localidad del cliente (por DNI) o en blanco
        out["LOCALID"] = [cliente.get("localidad", "") for cliente in clientes]

        # CUENTA - Cuenta contable
        out["CUENTA"] = [cuenta_contable_default] * len(clientes)
        
        return out[columnas_xubio]

    def verificar_compatibilidad_columnas(self, resultado_validacion: dict) -> dict:
        """Verifica la compatibilidad de columnas entre archivos"""
        compatibilidad = {
            "estado": "OK",
            "problemas": [],
            "recomendaciones": []
        }
        
        # Verificar archivo portal
        if "portal" in resultado_validacion and resultado_validacion["portal"]["estado"] == "OK":
            columnas_portal = resultado_validacion["portal"]["columnas"]
            
            # Verificar columnas cr√≠ticas
            tipo_doc_encontrado = self._encontrar_columna(columnas_portal, ['tipo_doc', 'tipo_documento', 'tipo', 'ct_kind0f', 'TIPO_DOC'])
            numero_doc_encontrado = self._encontrar_columna(columnas_portal, ['Numero de Documento', 'numero de documento', 'numero_documento', 'nro. doc. comprador', 'nro doc comprador', 'nro. doc comprador', 'dni', 'cuit', 'CUIT', 'NUMERO_DOC'])
            nombre_encontrado = self._encontrar_columna(columnas_portal, ['nombre', 'razon_social', 'cliente', 'NOMBRE'])
            
            if not tipo_doc_encontrado:
                compatibilidad["estado"] = "ADVERTENCIA"
                compatibilidad["problemas"].append("Portal: No se encontr√≥ columna de tipo de documento")
                compatibilidad["recomendaciones"].append("Agregar columna 'TIPO_DOC' o 'tipo_documento'")
            
            if not numero_doc_encontrado:
                compatibilidad["estado"] = "ADVERTENCIA"
                compatibilidad["problemas"].append("Portal: No se encontr√≥ columna de n√∫mero de documento")
                compatibilidad["recomendaciones"].append("Agregar columna 'NUMERO_DOC' o 'numero_documento'")
            
            if not nombre_encontrado:
                compatibilidad["estado"] = "ADVERTENCIA"
                compatibilidad["problemas"].append("Portal: No se encontr√≥ columna de nombre")
                compatibilidad["recomendaciones"].append("Agregar columna 'NOMBRE' o 'nombre'")
            
            # Verificar formato de tipo de documento
            if tipo_doc_encontrado:
                muestra_tipo_doc = resultado_validacion["portal"]["muestra"][0].get(tipo_doc_encontrado, "")
                if str(muestra_tipo_doc) not in ['80', '96']:
                    compatibilidad["recomendaciones"].append(f"Portal: El valor '{muestra_tipo_doc}' en columna '{tipo_doc_encontrado}' no es un c√≥digo v√°lido (80=CUIT, 96=DNI)")
        
        # Verificar archivo Xubio
        if "xubio" in resultado_validacion and resultado_validacion["xubio"]["estado"] == "OK":
            columnas_xubio = resultado_validacion["xubio"]["columnas"]
            
            id_encontrado = self._encontrar_columna(columnas_xubio, ['cuit', 'dni', 'documento', 'identificador', 'numeroidentificacion', 'NUMEROIDENTIFICACION'])
            nombre_encontrado = self._encontrar_columna(columnas_xubio, ['nombre', 'razon', 'cliente', 'NOMBRE'])
            
            if not id_encontrado:
                compatibilidad["estado"] = "ADVERTENCIA"
                compatibilidad["problemas"].append("Xubio: No se encontr√≥ columna de identificador")
                compatibilidad["recomendaciones"].append("Agregar columna 'NUMEROIDENTIFICACION' o 'identificador'")
            
            if not nombre_encontrado:
                compatibilidad["estado"] = "ADVERTENCIA"
                compatibilidad["problemas"].append("Xubio: No se encontr√≥ columna de nombre")
                compatibilidad["recomendaciones"].append("Agregar columna 'NOMBRE' o 'nombre'")
        
        # Verificar archivo cliente (opcional)
        if "cliente" in resultado_validacion and resultado_validacion["cliente"]["estado"] == "OK":
            columnas_cliente = resultado_validacion["cliente"]["columnas"]
            
            provincia_encontrada = self._encontrar_columna(columnas_cliente, ['provincia', 'prov', 'Provincia / Estado / Region'])
            if not provincia_encontrada:
                compatibilidad["recomendaciones"].append("Cliente: No se encontr√≥ columna de provincia (opcional pero recomendado)")
        
        # Resumen final
        if compatibilidad["estado"] == "OK" and not compatibilidad["problemas"]:
            compatibilidad["mensaje"] = "‚úÖ Archivos compatibles y listos para procesar"
        elif compatibilidad["estado"] == "ADVERTENCIA":
            compatibilidad["mensaje"] = "‚ö†Ô∏è Archivos procesables pero con advertencias"
        else:
            compatibilidad["mensaje"] = "‚ùå Archivos con problemas cr√≠ticos"
        
        return compatibilidad
